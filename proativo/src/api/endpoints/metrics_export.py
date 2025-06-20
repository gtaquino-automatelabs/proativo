"""
Endpoints para exportação de métricas do sistema PROAtivo.

Este módulo implementa endpoints para exportar métricas do sistema
em diferentes formatos para análise externa e integração com ferramentas de BI.
"""

import io
import csv
import json
import logging
from datetime import datetime, timedelta
from typing import Dict, Any, List, Optional
from enum import Enum

from fastapi import APIRouter, Depends, HTTPException, Query, Response
from fastapi.responses import StreamingResponse
from sqlalchemy.ext.asyncio import AsyncSession

from ..dependencies import get_database_session, get_current_config, get_llm_service, get_cache_service, get_fallback_service
from ..config import Settings
from ...utils.error_handlers import ValidationError, DataProcessingError

# Configurar logging
logger = logging.getLogger(__name__)

# Criar router para endpoints de exportação
router = APIRouter(prefix="/export", tags=["metrics-export"])


class ExportFormat(str, Enum):
    """Formatos suportados para exportação."""
    JSON = "json"
    CSV = "csv"
    XLSX = "xlsx"


class MetricsScope(str, Enum):
    """Escopo das métricas para exportação."""
    ALL = "all"
    LLM = "llm"
    CACHE = "cache"
    FEEDBACK = "feedback"
    FALLBACK = "fallback"
    PERFORMANCE = "performance"
    UNKNOWN_QUERIES = "unknown_queries"


@router.get("/metrics")
async def export_metrics(
    format: ExportFormat = Query(ExportFormat.JSON, description="Formato de exportação"),
    scope: MetricsScope = Query(MetricsScope.ALL, description="Escopo das métricas"),
    include_historical: bool = Query(False, description="Incluir dados históricos"),
    days_back: int = Query(7, description="Dias para trás (se histórico)", ge=1, le=365),
    settings: Settings = Depends(get_current_config),
    llm_service=Depends(get_llm_service),
    cache_service=Depends(get_cache_service),
    fallback_service=Depends(get_fallback_service)
) -> Response:
    """
    Exporta métricas do sistema em formato especificado.
    
    Args:
        format: Formato de exportação (JSON, CSV, XLSX)
        scope: Escopo das métricas a exportar
        include_historical: Se deve incluir dados históricos
        days_back: Quantos dias para trás incluir
        settings: Configurações da aplicação
        llm_service: Serviço LLM
        cache_service: Serviço de cache
        fallback_service: Serviço de fallback
        
    Returns:
        Response: Arquivo com métricas exportadas
        
    Raises:
        HTTPException: Erro na exportação
    """
    try:
        logger.info(f"Iniciando exportação de métricas - Formato: {format}, Escopo: {scope}")
        
        # Coletar métricas baseadas no escopo
        metrics_data = await _collect_metrics_by_scope(
            scope, llm_service, cache_service, fallback_service
        )
        
        # Adicionar metadados da exportação
        export_metadata = {
            "export_timestamp": datetime.now().isoformat(),
            "export_format": format.value,
            "export_scope": scope.value,
            "include_historical": include_historical,
            "days_back": days_back if include_historical else 0,
            "system_version": "PROAtivo v1.0",
            "exported_by": "metrics_export_api"
        }
        
        # Estrutura final dos dados
        export_data = {
            "metadata": export_metadata,
            "metrics": metrics_data,
            "summary": _generate_metrics_summary(metrics_data)
        }
        
        # Exportar no formato solicitado
        if format == ExportFormat.JSON:
            return _export_json(export_data)
        elif format == ExportFormat.CSV:
            return _export_csv(export_data)
        elif format == ExportFormat.XLSX:
            return _export_xlsx(export_data)
        else:
            raise HTTPException(status_code=400, detail=f"Formato não suportado: {format}")
            
    except ValidationError as e:
        logger.warning(f"Erro de validação na exportação: {str(e)}")
        raise HTTPException(status_code=400, detail=f"Erro de validação: {str(e)}")
    
    except Exception as e:
        logger.error(f"Erro inesperado na exportação: {str(e)}")
        raise HTTPException(
            status_code=500,
            detail="Erro interno durante exportação de métricas"
        )


async def _collect_metrics_by_scope(
    scope: MetricsScope, 
    llm_service, 
    cache_service, 
    fallback_service
) -> Dict[str, Any]:
    """
    Coleta métricas baseadas no escopo especificado.
    
    Args:
        scope: Escopo das métricas
        llm_service: Serviço LLM
        cache_service: Serviço de cache
        fallback_service: Serviço de fallback
        
    Returns:
        Dict com métricas coletadas
    """
    metrics_data = {}
    
    # LLM Metrics
    if scope in [MetricsScope.ALL, MetricsScope.LLM]:
        if llm_service:
            try:
                llm_metrics = await llm_service.get_metrics()
                metrics_data["llm"] = llm_metrics
            except Exception as e:
                logger.warning(f"Erro ao coletar métricas LLM: {e}")
                metrics_data["llm"] = {"error": "Métricas não disponíveis"}
    
    # Cache Metrics
    if scope in [MetricsScope.ALL, MetricsScope.CACHE]:
        if cache_service:
            try:
                cache_metrics = await cache_service.get_metrics()
                metrics_data["cache"] = {
                    "total_requests": cache_metrics.total_requests,
                    "cache_hits": cache_metrics.cache_hits,
                    "cache_misses": cache_metrics.cache_misses,
                    "hit_rate": cache_metrics.hit_rate,
                    "miss_rate": cache_metrics.miss_rate,
                    "cache_size": cache_metrics.cache_size,
                    "max_cache_size": cache_metrics.max_cache_size,
                    "memory_usage_mb": cache_metrics.memory_usage_mb,
                    "average_response_size": cache_metrics.average_response_size,
                    "expired_entries": cache_metrics.expired_entries,
                    "stale_entries": cache_metrics.stale_entries
                }
            except Exception as e:
                logger.warning(f"Erro ao coletar métricas de cache: {e}")
                metrics_data["cache"] = {"error": "Métricas não disponíveis"}
    
    # Fallback Metrics
    if scope in [MetricsScope.ALL, MetricsScope.FALLBACK]:
        if fallback_service:
            try:
                fallback_metrics = fallback_service.get_metrics()
                metrics_data["fallback"] = {
                    "total_fallbacks": fallback_metrics.total_fallbacks,
                    "success_rate": fallback_metrics.success_rate,
                    "user_satisfaction": fallback_metrics.user_satisfaction,
                    "fallbacks_by_trigger": fallback_metrics.fallbacks_by_trigger,
                    "fallbacks_by_strategy": fallback_metrics.fallbacks_by_strategy
                }
            except Exception as e:
                logger.warning(f"Erro ao coletar métricas de fallback: {e}")
                metrics_data["fallback"] = {"error": "Métricas não disponíveis"}
    
    # Unknown Queries específicas
    if scope in [MetricsScope.ALL, MetricsScope.UNKNOWN_QUERIES]:
        if llm_service and hasattr(llm_service, 'unknown_query_count'):
            metrics_data["unknown_queries"] = {
                "total_unknown_queries": llm_service.unknown_query_count,
                "categories": llm_service.unknown_query_categories,
                "unknown_query_rate": (llm_service.unknown_query_count / llm_service.request_count) if llm_service.request_count > 0 else 0
            }
    
    return metrics_data


def _generate_metrics_summary(metrics_data: Dict[str, Any]) -> Dict[str, Any]:
    """
    Gera resumo das métricas coletadas.
    
    Args:
        metrics_data: Dados das métricas
        
    Returns:
        Dict com resumo das métricas
    """
    summary = {
        "total_services": len([k for k in metrics_data.keys() if not metrics_data[k].get("error")]),
        "services_with_errors": len([k for k in metrics_data.keys() if metrics_data[k].get("error")]),
        "collection_timestamp": datetime.now().isoformat(),
        "health_indicators": {}
    }
    
    # Indicadores de saúde
    if "llm" in metrics_data and not metrics_data["llm"].get("error"):
        llm_data = metrics_data["llm"]
        summary["health_indicators"]["llm_health"] = {
            "error_rate": llm_data.get("error_rate", 0),
            "fallback_rate": llm_data.get("fallback_rate", 0),
            "status": "healthy" if llm_data.get("error_rate", 0) < 0.1 else "warning"
        }
    
    if "cache" in metrics_data and not metrics_data["cache"].get("error"):
        cache_data = metrics_data["cache"]
        summary["health_indicators"]["cache_health"] = {
            "hit_rate": cache_data.get("hit_rate", 0),
            "memory_usage": cache_data.get("memory_usage_mb", 0),
            "status": "healthy" if cache_data.get("hit_rate", 0) > 0.3 else "warning"
        }
    
    return summary


def _export_json(data: Dict[str, Any]) -> Response:
    """
    Exporta dados em formato JSON.
    
    Args:
        data: Dados para exportar
        
    Returns:
        Response: Resposta JSON
    """
    json_str = json.dumps(data, indent=2, ensure_ascii=False, default=str)
    filename = f"proativo_metrics_{datetime.now().strftime('%Y%m%d_%H%M%S')}.json"
    
    return Response(
        content=json_str,
        media_type="application/json",
        headers={"Content-Disposition": f"attachment; filename={filename}"}
    )


def _export_csv(data: Dict[str, Any]) -> StreamingResponse:
    """
    Exporta dados em formato CSV.
    
    Args:
        data: Dados para exportar
        
    Returns:
        StreamingResponse: Resposta CSV
    """
    output = io.StringIO()
    writer = csv.writer(output)
    
    # Cabeçalhos
    writer.writerow(["Categoria", "Métrica", "Valor", "Timestamp"])
    
    # Metadados
    metadata = data.get("metadata", {})
    for key, value in metadata.items():
        writer.writerow(["metadata", key, value, metadata.get("export_timestamp", "")])
    
    # Métricas por categoria
    metrics = data.get("metrics", {})
    timestamp = metadata.get("export_timestamp", "")
    
    for category, category_data in metrics.items():
        if isinstance(category_data, dict) and not category_data.get("error"):
            _write_dict_to_csv(writer, category, category_data, timestamp)
    
    # Resumo
    summary = data.get("summary", {})
    for key, value in summary.items():
        if key != "health_indicators":
            writer.writerow(["summary", key, value, timestamp])
    
    output.seek(0)
    filename = f"proativo_metrics_{datetime.now().strftime('%Y%m%d_%H%M%S')}.csv"
    
    return StreamingResponse(
        io.StringIO(output.getvalue()),
        media_type="text/csv",
        headers={"Content-Disposition": f"attachment; filename={filename}"}
    )


def _write_dict_to_csv(writer, category: str, data: Dict[str, Any], timestamp: str, prefix: str = ""):
    """
    Escreve dicionário no CSV de forma recursiva.
    
    Args:
        writer: Writer CSV
        category: Categoria da métrica
        data: Dados do dicionário
        timestamp: Timestamp da exportação
        prefix: Prefixo para métricas aninhadas
    """
    for key, value in data.items():
        full_key = f"{prefix}.{key}" if prefix else key
        
        if isinstance(value, dict):
            _write_dict_to_csv(writer, category, value, timestamp, full_key)
        else:
            writer.writerow([category, full_key, value, timestamp])


def _export_xlsx(data: Dict[str, Any]) -> Response:
    """
    Exporta dados em formato XLSX.
    
    Args:
        data: Dados para exportar
        
    Returns:
        Response: Resposta XLSX
    """
    try:
        import openpyxl
        from openpyxl.styles import Font, PatternFill
        
        wb = openpyxl.Workbook()
        
        # Aba de resumo
        ws_summary = wb.active
        ws_summary.title = "Resumo"
        
        # Headers
        ws_summary.cell(1, 1, "PROAtivo - Exportação de Métricas").font = Font(bold=True, size=14)
        ws_summary.cell(2, 1, f"Gerado em: {data['metadata']['export_timestamp']}")
        
        row = 4
        summary = data.get("summary", {})
        for key, value in summary.items():
            ws_summary.cell(row, 1, key).font = Font(bold=True)
            ws_summary.cell(row, 2, str(value))
            row += 1
        
        # Aba para cada categoria de métrica
        metrics = data.get("metrics", {})
        for category, category_data in metrics.items():
            if isinstance(category_data, dict) and not category_data.get("error"):
                ws = wb.create_sheet(title=category.capitalize())
                _write_dict_to_excel_sheet(ws, category_data)
        
        # Salvar em buffer
        buffer = io.BytesIO()
        wb.save(buffer)
        buffer.seek(0)
        
        filename = f"proativo_metrics_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
        
        return Response(
            content=buffer.getvalue(),
            media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            headers={"Content-Disposition": f"attachment; filename={filename}"}
        )
        
    except ImportError:
        logger.warning("openpyxl não disponível, retornando CSV")
        return _export_csv(data)


def _write_dict_to_excel_sheet(worksheet, data: Dict[str, Any], start_row: int = 1):
    """
    Escreve dicionário em planilha Excel.
    
    Args:
        worksheet: Planilha do Excel
        data: Dados do dicionário
        start_row: Linha inicial
    """
    row = start_row
    
    # Headers
    worksheet.cell(row, 1, "Métrica").font = Font(bold=True)
    worksheet.cell(row, 2, "Valor").font = Font(bold=True)
    row += 1
    
    for key, value in data.items():
        if isinstance(value, dict):
            worksheet.cell(row, 1, f"=== {key.upper()} ===").font = Font(bold=True)
            row += 1
            for sub_key, sub_value in value.items():
                worksheet.cell(row, 1, f"  {sub_key}")
                worksheet.cell(row, 2, str(sub_value))
                row += 1
        else:
            worksheet.cell(row, 1, key)
            worksheet.cell(row, 2, str(value))
            row += 1


@router.get("/health")
async def export_health_check() -> Dict[str, Any]:
    """
    Endpoint de health check para o sistema de exportação.
    
    Returns:
        Dict com status do sistema de exportação
    """
    try:
        # Verificar dependências opcionais
        dependencies = {
            "openpyxl": False,
            "csv": True,  # Built-in
            "json": True  # Built-in
        }
        
        try:
            import openpyxl
            dependencies["openpyxl"] = True
        except ImportError:
            pass
        
        return {
            "status": "healthy",
            "service": "metrics_export",
            "supported_formats": ["json", "csv"] + (["xlsx"] if dependencies["openpyxl"] else []),
            "dependencies": dependencies,
            "timestamp": datetime.now().isoformat()
        }
        
    except Exception as e:
        logger.error(f"Erro no health check de exportação: {str(e)}")
        return {
            "status": "error",
            "error": str(e),
            "timestamp": datetime.now().isoformat()
        }


@router.get("/formats")
async def get_supported_formats() -> Dict[str, Any]:
    """
    Retorna formatos de exportação suportados.
    
    Returns:
        Dict com formatos suportados
    """
    formats = {
        "json": {
            "description": "JavaScript Object Notation",
            "extension": ".json",
            "media_type": "application/json",
            "supported": True
        },
        "csv": {
            "description": "Comma Separated Values",
            "extension": ".csv", 
            "media_type": "text/csv",
            "supported": True
        },
        "xlsx": {
            "description": "Microsoft Excel Spreadsheet",
            "extension": ".xlsx",
            "media_type": "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            "supported": False
        }
    }
    
    # Verificar se XLSX é suportado
    try:
        import openpyxl
        formats["xlsx"]["supported"] = True
    except ImportError:
        pass
    
    return {
        "supported_formats": formats,
        "scopes": [scope.value for scope in MetricsScope],
        "timestamp": datetime.now().isoformat()
    } 