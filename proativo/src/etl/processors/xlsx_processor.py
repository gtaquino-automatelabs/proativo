"""
Processador para arquivos XLSX usando OpenPyXL.

Responsável por ler, validar e transformar dados de equipamentos e manutenções
a partir de arquivos Excel (.xlsx), preparando-os para inserção no banco de dados.
"""

import logging
from typing import List, Dict, Any, Optional
from datetime import datetime
from pathlib import Path
import re

try:
    from openpyxl import load_workbook
except ImportError:
    raise ImportError("OpenPyXL não está instalado. Execute: pip install openpyxl")

from ..exceptions import DataProcessingError, FileFormatError

logger = logging.getLogger(__name__)


class XLSXProcessor:
    """Processador para arquivos XLSX com suporte a múltiplas planilhas."""
    
    def __init__(self):
        """Inicializa o processador XLSX."""
        pass
    
    def _infer_equipment_type(self, code: str, name: str) -> str:
        """Infere o tipo de equipamento baseado no código e nome.
        
        Args:
            code: Código do equipamento
            name: Nome do equipamento
            
        Returns:
            Tipo inferido do equipamento
        """
        # Combina código e nome para análise
        text_to_analyze = f"{code} {name}".lower()
        
        # Mapeamento de palavras-chave para tipos
        type_keywords = {
            'Transformador': ['transformador', 'tr-', 'trafo'],
            'Disjuntor': ['disjuntor', 'dj-', 'breaker'],
            'Seccionadora': ['seccionadora', 'sc-', 'chave', 'switch'],
            'Para-raios': ['para-raios', 'pr-', 'pararaios', 'surge'],
            'Religador': ['religador', 're-', 'recloser'],
            'Capacitor': ['capacitor', 'cp-', 'banco'],
            'Regulador': ['regulador', 'rg-', 'regulator'],
            'Medidor': ['medidor', 'md-', 'meter'],
        }
        
        # Procura por palavras-chave
        for equipment_type, keywords in type_keywords.items():
            for keyword in keywords:
                if keyword in text_to_analyze:
                    return equipment_type
        
        # Fallback baseado em código
        if code:
            code_upper = code.upper()
            if code_upper.startswith('TR-'):
                return 'Transformador'
            elif code_upper.startswith('DJ-'):
                return 'Disjuntor'
            elif code_upper.startswith('SC-'):
                return 'Seccionadora'
            elif code_upper.startswith('PR-'):
                return 'Para-raios'
            elif code_upper.startswith('RE-'):
                return 'Religador'
            elif code_upper.startswith('CP-'):
                return 'Capacitor'
            elif code_upper.startswith('RG-'):
                return 'Regulador'
            elif code_upper.startswith('MD-'):
                return 'Medidor'
        
        # Última tentativa: fallback genérico
        return 'Equipamento'
        
    def process_equipment_xlsx(self, file_path: Path, sheet_name: str = None) -> List[Dict[str, Any]]:
        """Processa arquivo XLSX de equipamentos.
        
        Args:
            file_path: Caminho para o arquivo
            sheet_name: Nome da planilha específica
            
        Returns:
            Lista de dicionários com dados de equipamentos
        """
        logger.info(f"Processando XLSX de equipamentos: {file_path}")
        
        try:
            workbook = load_workbook(file_path, read_only=True, data_only=True)
            
            # Seleciona planilha
            if sheet_name and sheet_name in workbook.sheetnames:
                worksheet = workbook[sheet_name]
            else:
                worksheet = workbook.active
            
            # Extrai cabeçalhos da primeira linha
            headers = []
            for col in range(1, worksheet.max_column + 1):
                cell = worksheet.cell(row=1, column=col)
                if cell.value:
                    header = str(cell.value).strip().lower()
                    header = re.sub(r'[^\w\s]', '', header)
                    header = re.sub(r'\s+', '_', header)
                    headers.append(header)
                else:
                    headers.append(f'column_{col}')
            
            # Mapeamento de colunas
            column_mapping = {
                'id': 'code',  # Mapeia campo 'id' do XLSX para 'code'
                'equipamento': 'code', 'codigo': 'code', 'codigo_equipamento': 'code',
                'nome': 'name', 'nome_equipamento': 'name', 'equipment_name': 'name',
                'descricao': 'description', 
                # Múltiplas formas de mapear equipment_type
                'tipo': 'equipment_type', 'type': 'equipment_type', 'tipo_equipamento': 'equipment_type',
                'equipment_type': 'equipment_type', 'category': 'equipment_type',
                'criticidade': 'criticality', 'localizacao': 'location', 'subestacao': 'substation',
                'fabricante': 'manufacturer', 'modelo': 'model', 'numero_serie': 'serial_number',
                'ano_fabricacao': 'manufacturing_year', 'data_instalacao': 'installation_date',
                'tensao_nominal': 'rated_voltage', 'potencia_nominal': 'rated_power',
                'corrente_nominal': 'rated_current', 'status': 'status'
            }
            
            # Extrai dados
            equipment_records = []
            for row_num in range(2, worksheet.max_row + 1):
                record = {}
                has_data = False
                
                for col_idx, header in enumerate(headers):
                    cell = worksheet.cell(row=row_num, column=col_idx + 1)
                    value = cell.value
                    
                    if value is not None:
                        has_data = True
                        # Mapeia para nome padrão
                        standard_name = column_mapping.get(header, header)
                        
                        # Converte tipos
                        if isinstance(value, datetime):
                            record[standard_name] = value
                        elif isinstance(value, (int, float)):
                            record[standard_name] = value
                        else:
                            str_value = str(value).strip()
                            if str_value:
                                record[standard_name] = str_value
                
                if has_data:
                    # Valida e garante campo 'code' obrigatório
                    if 'code' not in record or not record['code']:
                        # Tenta usar outros campos como fallback
                        fallback_code = None
                        for field in ['id', 'equipment_id', 'equipment_code', 'codigo', 'equipamento']:
                            if field in record and record[field]:
                                fallback_code = str(record[field]).strip().upper()
                                break
                        
                        if fallback_code:
                            record['code'] = fallback_code
                            logger.warning(f"Campo 'code' ausente, usando fallback: {fallback_code}")
                        else:
                            # Gera código sequencial se nenhum campo disponível
                            record['code'] = f"EQUIP-{len(equipment_records)+1:03d}"
                            logger.warning(f"Campo 'code' ausente, gerado código: {record['code']}")
                    
                    # Valida e garante campo 'equipment_type' obrigatório
                    if 'equipment_type' not in record or not record['equipment_type']:
                        # Infere o tipo baseado no código e nome
                        code = record.get('code', '')
                        name = record.get('name', '')
                        inferred_type = self._infer_equipment_type(code, name)
                        record['equipment_type'] = inferred_type
                        logger.info(f"Campo 'equipment_type' ausente para {code}, inferido: {inferred_type}")
                    
                    record['metadata_json'] = {
                        'source_file': file_path.name,
                        'source_format': 'XLSX',
                        'source_sheet': worksheet.title,
                        'processed_at': datetime.now().isoformat(),
                        'inferred_fields': []
                    }
                    
                    # Registra campos inferidos nos metadados
                    if 'equipment_type' in record:
                        record['metadata_json']['inferred_fields'].append('equipment_type')
                    
                    equipment_records.append(record)
            
            workbook.close()
            logger.info(f"Processados {len(equipment_records)} equipamentos do XLSX")
            return equipment_records
            
        except Exception as e:
            raise DataProcessingError(f"Erro ao processar XLSX de equipamentos {file_path}: {str(e)}")
    
    def process_maintenance_xlsx(self, file_path: Path, sheet_name: str = None) -> List[Dict[str, Any]]:
        """Processa arquivo XLSX de manutenções.
        
        Args:
            file_path: Caminho para o arquivo
            sheet_name: Nome da planilha específica
            
        Returns:
            Lista de dicionários com dados de manutenções
        """
        logger.info(f"Processando XLSX de manutenções: {file_path}")
        
        try:
            workbook = load_workbook(file_path, read_only=True, data_only=True)
            
            # Seleciona planilha
            if sheet_name and sheet_name in workbook.sheetnames:
                worksheet = workbook[sheet_name]
            else:
                worksheet = workbook.active
            
            # Extrai cabeçalhos da primeira linha
            headers = []
            for col in range(1, worksheet.max_column + 1):
                cell = worksheet.cell(row=1, column=col)
                if cell.value:
                    header = str(cell.value).strip().lower()
                    header = re.sub(r'[^\w\s]', '', header)
                    header = re.sub(r'\s+', '_', header)
                    headers.append(header)
                else:
                    headers.append(f'column_{col}')
            
            # Mapeamento de colunas
            column_mapping = {
                'id': 'maintenance_code',  # Mapeia 'id' para 'maintenance_code'
                'equipment_id': 'equipment_id',  # Mantém equipment_id  
                'equipamento_id': 'equipment_id',  # Mapeia equipamento_id
                'codigo_equipamento': 'equipment_id',  # Mapeia codigo_equipamento
                'codigo_manutencao': 'maintenance_code', 'tipo_manutencao': 'maintenance_type',
                'prioridade': 'priority', 'titulo': 'title', 'descricao': 'description',
                'trabalho_realizado': 'work_performed', 'data_programada': 'scheduled_date',
                'data_inicio': 'start_date', 'data_conclusao': 'completion_date',
                'duracao_horas': 'duration_hours', 'resultado': 'result',
                'tecnico': 'technician', 'equipe': 'team', 'contratada': 'contractor',
                'custo_estimado': 'estimated_cost', 'custo_real': 'actual_cost',
                'observacoes': 'observations'
            }
            
            # Extrai dados
            maintenance_records = []
            for row_num in range(2, worksheet.max_row + 1):
                record = {}
                has_data = False
                
                for col_idx, header in enumerate(headers):
                    cell = worksheet.cell(row=row_num, column=col_idx + 1)
                    value = cell.value
                    
                    if value is not None:
                        has_data = True
                        # Mapeia para nome padrão
                        standard_name = column_mapping.get(header, header)
                        
                        # Converte tipos
                        if isinstance(value, datetime):
                            record[standard_name] = value
                        elif isinstance(value, (int, float)):
                            record[standard_name] = value
                        else:
                            str_value = str(value).strip()
                            if str_value:
                                record[standard_name] = str_value
                
                if has_data:
                    record['metadata_json'] = {
                        'source_file': file_path.name,
                        'source_format': 'XLSX',
                        'source_sheet': worksheet.title,
                        'processed_at': datetime.now().isoformat()
                    }
                    maintenance_records.append(record)
            
            workbook.close()
            logger.info(f"Processadas {len(maintenance_records)} manutenções do XLSX")
            return maintenance_records
            
        except Exception as e:
            raise DataProcessingError(f"Erro ao processar XLSX de manutenções {file_path}: {str(e)}") 